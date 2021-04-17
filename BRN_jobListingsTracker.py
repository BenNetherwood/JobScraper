import urllib
import requests
from bs4 import BeautifulSoup
import selenium
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
import pandas as pd
import os
import openpyxl
from datetime import datetime, timedelta, date
from bs4.element import Comment
import urllib.request
from wordcloud import WordCloud, STOPWORDS 
import matplotlib.pyplot as plt 
import re
today = date.today()
import xlrd
import xlsxwriter
import numpy as np

## import last run of the codes 


try:
    f = open('results.xlsx')
    f.close()
except FileNotFoundError:
    
    workbook = xlsxwriter.Workbook('results.xlsx')
    
    # The workbook object is then used to add new 
    # worksheet via the add_worksheet() method.
    worksheet = workbook.add_worksheet()
    
    # Use the worksheet object to write
    # data via the write() method.
    
    # Finally, close the Excel file
    # via the close() method.
    workbook.close()
    


salariesListLOWER=[]
salariesListUPPER=[]

ps = openpyxl.load_workbook("results.xlsx")
sheet = ps["Sheet1"]





def find_jobs_from(website, job_title, location, desired_characs, filename="results.xlsx"):    
    """
    This function extracts all the desired characteristics of all new job postings
    of the title and location specified and returns them in single file.
    The arguments it takes are:
        - Website: to specify which website to search (options: 'Indeed' or 'CWjobs')
        - Job_title
        - Location
        - Desired_characs: this is a list of the job characteristics of interest,
            from titles, companies, links and date_listed.
        - Filename: to specify the filename and format of the output.
            Default is .xls file called 'results.xls'
    """
    
    if website == 'Indeed':
        job_soup = load_indeed_jobs_div(job_title, location)
        jobs_list, num_listings,newjobscouunt = extract_job_information_indeed(job_soup, desired_characs)
    

    
    save_jobs_to_excel(jobs_list, filename)
 
    print('{} new job postings retrieved.'.format(newjobscouunt, 
                                                                        ))
    print('{} jobs in total in database..'.format(num_listings, 
                                                                        ))
    

## ======================= GENERIC FUNCTIONS ======================= ##

def save_jobs_to_excel(jobs_list, filename):
    jobs = pd.DataFrame(jobs_list)
    jobs.set_index('titles', inplace=True)
    jobs=jobs.drop_duplicates('links',keep='first')
    
    jobs.to_excel(filename)

    # making data frame from csv file 








## ================== FUNCTIONS FOR INDEED.CO.UK =================== ##

def load_indeed_jobs_div(job_title, location):
    getVars = {'q' : job_title, 'l' : location, 'fromage' : 'last', 'sort' : 'date'}
    url = ('https://www.indeed.co.uk/jobs?' + urllib.parse.urlencode(getVars))
    page = requests.get(url)
    soup = BeautifulSoup(page.content, "html.parser")
    job_soup = soup.find(id="resultsCol")



    return job_soup

def extract_job_information_indeed(job_soup, desired_characs):

    job_elems = job_soup.find_all('div', class_='jobsearch-SerpJobCard')
     
    cols = []
    exceladdedpriorindexes=[]
    jobs_list = {}
    index=0

    cols.append('titles')
    cols.append('companies')
    cols.append('links')
    cols.append('date_listed')
    cols.append('salary')
    cols.append('desc')

    extracted_infotitles = []
    extracted_infolink = []
    extracted_infocompany = []
    extracted_infodate = []
    extracted_infodesc = []
    extracted_infosalary=[]

    NoveltyIndexes=[]
    NOVEL = np.array([])
    foundindexList=[]
    for job_elem in job_elems:
        index=index+1


        extracted_info_Current = []
        if 'titles' in desired_characs:
            titles = []
            
            
            titles.append(extract_job_title_indeed(job_elem))
            titles = str(titles)[2:-2]
            joblisting=titles
            if joblisting[(len(joblisting)-2):len(joblisting)]=='\\n':
                joblisting=joblisting[0:len(joblisting)-1]

            if joblisting[(len(joblisting)-1):len(joblisting)]=='\\':
                joblisting=joblisting[0:len(joblisting)-1]  

            extracted_infotitles.append(joblisting)    
            extracted_info_Current.append(joblisting)                 
        
        if 'companies' in desired_characs:
            companies = []
            
        
            companies.append(extract_company_indeed(job_elem))
            companies = str(companies)[2:-2]
            extracted_infocompany.append(companies)
            extracted_info_Current.append(companies)      
        
        if 'links' in desired_characs:
            links = []
         
            linkreal,desc=extract_link_indeed(job_elem)
            desc=desc[304:len(desc)-261] #remove generic and padder
            if index==1:
                MasterJobDescs=desc
            else:
                MasterJobDescs=MasterJobDescs+' ' +desc
            
            salaryfind=findOccurrences(desc, '£')
            spacesfind=findOccurrences(desc, ' ')
            stopsfind=findOccurrences(desc, '.')

            if len(salaryfind)==0 or len(salaryfind)==1 :
                stringgoal=-1
            else:
                bEND=0
                bleftdigitsfirst=0
                bsecondnumber=0
                bspaceender=-1
                ender=-1

                for x in range(1,15):

                    if bEND==0:
                        character=desc[salaryfind[1]+x]

                        bDIGIT=str.isdigit(character)
                        bspace=0
                        bstop=0
                        bdash=0
                        if character==' ':
                            bspace=1
                        if character=='.':
                            bstop=1
                        if character=='-':
                            bdash=1

                        

                        if bleftdigitsfirst==1 and (bstop or bspace) and bsecondnumber:
                            ender=x
                            bEND=1

                        if bleftdigitsfirst==1 and bDIGIT==1:
                            bsecondnumber=1
                        if (bspace or bdash) and bleftdigitsfirst==0:

                            bleftdigitsfirst=1
                            bspaceender=x

                        
                        

                if ender==-1:
                    stringgoal=desc[salaryfind[1]+1:salaryfind[1]+bspaceender]

                else:
                    stringgoal=desc[salaryfind[1]:salaryfind[1]+ender]
                
                stringgoal=re.sub(",","",stringgoal)
                stringgoal = [stringgoal.replace(".", "") for stringgoal in stringgoal]
                rebuildstr=''
                stri=-1
                for stringgoalcu in stringgoal:
                    stri=stri+1
                    a=stringgoal[stri]
                    rebuildstr=rebuildstr+a
                stringgoal=rebuildstr
                if stringgoal[len(stringgoal)-1]=='K' or stringgoal[len(stringgoal)-1]=='k':
                    stringgoal=stringgoal[0:len(stringgoal)-1]
                    
                try: 
                    stringgoal=int(stringgoal)
                except Exception:
                    stringgoal=-1
                    
                spacesfind=findOccurrences(desc, ' ')
                stopsfind=findOccurrences(desc, '-')
                
                if stringgoal>150 and stringgoal<600:
                    stringgoal=stringgoal*52.3 # assume weekly (converto yearly approx rate for comparable)...

                if stringgoal<150:
                    stringgoal=stringgoal*1000 # is in K therefore assumed

            salariesListUPPER.append(stringgoal)
            extracted_infosalary.append(stringgoal)
            extracted_infodesc.append(desc)    
            
            


            links.append(linkreal)
            links = str(links)[2:-2]
            extracted_infolink.append(links)
            extracted_info_Current.append(links)      
        
        if 'date_listed' in desired_characs:
            dates = []
           
            
            

            dates.append(extract_date_indeed(job_elem))
            # conver to real date 
            if dates[0]=='Today' or dates[0]=='Just posted' :

                dates = today.strftime("%d/%m/%Y")

            else:
                daysago=dates[0]
                daysago=int(daysago[0:(len(daysago)-8)])
                current_date = today
                new_date = current_date - timedelta(days=daysago)
                dates = new_date.strftime("%d/%m/%Y")

            extracted_infodate.append(dates)
            

            extracted_info_Current.append(dates)      
        
       

         
        linkreal,desc=extract_link_indeed(job_elem)
        links = str(links)[2:-2]
        #extracted_infolink.append(links)
        extracted_info_Current.append(links)  
        
        foundindex=-1
        bfoundThisEXCEL=-1
        bNEW=1
        for row in range(2, sheet.max_row+1):
            bfoundThisEXCEL=bfoundThisEXCEL+1
            
            joblisting = sheet["A" + str(row)].value
            cmpny = sheet["B" + str(row)].value
            lnk = sheet["C" + str(row)].value
            date = sheet["D" + str(row)].value
            desc = sheet["F" + str(row)].value
            salary = sheet["E" + str(row)].value
            valuer=-1


            if joblisting==extracted_info_Current[0] and cmpny==extracted_info_Current[1] and extracted_info_Current[2]==lnk or  extracted_info_Current[3]==date:
                bNEW=0
                foundindex=bfoundThisEXCEL


        NoveltyIndexes.append(bNEW)
        NOVEL= np.append(NOVEL, bNEW)

        jobs_list[cols[0]]=(extracted_infotitles)
        jobs_list[cols[1]]=(extracted_infocompany)
        jobs_list[cols[2]]=(extracted_infolink)
        jobs_list[cols[3]]=(extracted_infodate)
        jobs_list[cols[4]]=(extracted_infosalary)
        jobs_list[cols[5]]=(extracted_infodesc)
        

        


        



    joblistingCURRENTLIST=jobs_list[cols[0]]
    DatesCURRENTLIST=jobs_list[cols[3]]
    COMPANYlistingCURRENTLIST=jobs_list[cols[1]]
    LINKlistingCURRENTLIST=jobs_list[cols[2]]
    bNEW=1
    for row in range(2, sheet.max_row+1):
        bfoundThisEXCEL=0
        foundindex=-1
        joblisting = sheet["B" + str(row)].value
        
        if joblisting[(len(joblisting)-2):len(joblisting)]=='\\n':
            joblisting=joblisting[0:len(joblisting)-1]

        if joblisting[(len(joblisting)-1):len(joblisting)]=='\\':
            joblisting=joblisting[0:len(joblisting)-1]    



        cmpny = sheet["C" + str(row)].value
        lnk = sheet["D" + str(row)].value
        date = sheet["E" + str(row)].value
        desc = sheet["F" + str(row)].value
        salary = sheet["G" + str(row)].value
        valuer=-1
        for valuerxxxx in joblistingCURRENTLIST:
            valuer=valuer+1
            extracted_info_Current[0]=joblistingCURRENTLIST[valuer]
            extracted_info_Current[1]=COMPANYlistingCURRENTLIST[valuer]
            extracted_info_Current[2]=LINKlistingCURRENTLIST[valuer]
            extracted_info_Current[3]=DatesCURRENTLIST[valuer]

            if joblisting==extracted_info_Current[0] and cmpny==extracted_info_Current[1] and extracted_info_Current[2]==lnk or  extracted_info_Current[3]==date:
                bNEW=0
                foundindex=row

        exceladdedpriorindexes.append(foundindex)

    
            




    
    SCANNER_NOVEL=np.where(NOVEL==1)
    flag =  np.any(NOVEL)
    if flag==1:
        num_NEW_listings = SCANNER_NOVEL.shape[0]
    else:
        num_NEW_listings = 0



    # append jobs not found in new search (upto 3 months ago)... 
    


    
    jobs_list[cols[0]]=(extracted_infotitles)
    jobs_list[cols[1]]=(extracted_infocompany)
    jobs_list[cols[2]]=(extracted_infolink)
    jobs_list[cols[3]]=(extracted_infodate)
    jobs_list[cols[4]]=(extracted_infosalary)
    jobs_list[cols[5]]=(extracted_infodesc)
     
    exceladdedpriorindexes
    indexer=-1
    for indexExcel in range(2, sheet.max_row+1):
        joblisting = sheet["A" + str(indexExcel)].value
        if joblisting[(len(joblisting)-2):len(joblisting)]=='\\n':
            joblisting=joblisting[0:len(joblisting)-1]

        if joblisting[(len(joblisting)-1):len(joblisting)]=='\\':
            joblisting=joblisting[0:len(joblisting)-1]  
        

        cmpny = sheet["B" + str(indexExcel)].value
        if cmpny=='Data Idols':
            her=0

        lnk = sheet["C" + str(indexExcel)].value
        date = sheet["D" + str(indexExcel)].value
        desc = sheet["F" + str(indexExcel)].value
        salary = sheet["E" + str(indexExcel)].value

        indexer=indexer+1
        
        addback=1
        

        if addback:

            extracted_infotitles.append(joblisting)  
            extracted_infocompany.append(cmpny)
            extracted_infolink.append(lnk)
            extracted_infodate.append(date)
            extracted_infosalary.append(salary)
            extracted_infodesc.append(desc)








    jobs_list[cols[0]]=(extracted_infotitles)
    jobs_list[cols[1]]=(extracted_infocompany)
    jobs_list[cols[2]]=(extracted_infolink)
    jobs_list[cols[3]]=(extracted_infodate)
    jobs_list[cols[4]]=(extracted_infosalary)
    jobs_list[cols[5]]=(extracted_infodesc)


    num_listings = len(jobs_list[cols[0]])

    sub_list = ["Indeed Events","Report job","Content","Indeed","Indeed Anti","Find","Indeed Privacy", "Indeed Home", "Now Apply","use", "Centres Cookies", "Employers Post", "Skip", "PHASTAR", "Centre Indeed","Report Job", "Employers Find", "Advice Browse","Info Follow","Indeed Blog","Jobs Browse","Browse Companies","Help Centre","days ago","career Advice","Job Search","Indeed Anti","Privacy Terms","site Save","Job Skip","Post Job","Cookies Privacy","Slavery Statement"]
    for sub in sub_list:
        MasterJobDescs = MasterJobDescs.replace(' ' + sub + ' ', ' ')

    stopwords = set(STOPWORDS) 
    wordcloud = WordCloud(width = 800, height = 800, 
    background_color ='white', 
    stopwords = stopwords, 
    min_font_size = 10).generate(MasterJobDescs) 

    # plot the WordCloud image                        
    plt.figure(figsize = (8, 8), facecolor = None) 
    plt.imshow(wordcloud) 
    plt.axis("off") 
    plt.tight_layout(pad = 0) 
            
    plt.show() 

    jobs = pd.DataFrame(jobs_list)
    jobs.set_index('titles', inplace=True)
    jobs=jobs.drop_duplicates('links',keep='first')
    
    



    return jobs_list, len(jobs.index), num_NEW_listings


def extract_job_title_indeed(job_elem):
    title_elem = job_elem.find('h2', class_='title')
    title = title_elem.text.strip()
    if title[len(title)-3:len(title)]=='new':

        title=title[0:len(title)-3]



    return title

def extract_company_indeed(job_elem):
    company_elem = job_elem.find('span', class_='company')
    company = company_elem.text.strip()
    return company

def extract_link_indeed(job_elem):
    link = job_elem.find('a')['id']
    link = 'https://uk.indeed.com/viewjob?jk=' + link[3:len(link)] + '&from=serp&vjs=3'

    html = urllib.request.urlopen(link).read()
    outputdesc=(text_from_html(html))



    return link, outputdesc

def findOccurrences(s, ch):
    return [i for i, letter in enumerate(s) if letter == ch]


def extract_date_indeed(job_elem):
    date_elem = job_elem.find('span', class_='date')
    date = date_elem.text.strip()
    return date


def tag_visible(element):
    if element.parent.name in ['style', 'script', 'head', 'title', 'meta', '[document]']:
        return False
    if isinstance(element, Comment):
        return False
    return True


def text_from_html(body):
    soup = BeautifulSoup(body, 'html.parser')
    texts = soup.findAll(text=True)
    visible_texts = filter(tag_visible, texts)  
    return u" ".join(t.strip() for t in visible_texts)    


    
desiredCh={'titles', 'companies', 'links', 'date_listed'}
find_jobs_from('Indeed', 'data scientist', 'Manchester', desiredCh)
#find_jobs_from('Indeed', 'machine learning', 'Manchester', desiredCh)

